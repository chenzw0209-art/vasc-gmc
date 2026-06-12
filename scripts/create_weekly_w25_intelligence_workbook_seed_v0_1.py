#!/usr/bin/env python3
"""Create the W25 seed intelligence workbook for lead collection."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


OUT = Path(r"Z:\增长分析中台\线索采集\售前情报库_W25（周更新）.xlsx")
WEEK = "W25"
SEARCH_DATE = "2026-06-12"


CUSTOMER_HEADERS = [
    "候选ID", "周次", "检索日期", "钛动分类", "一级行业", "二级行业", "三级品类",
    "应用/商品名称", "背后公司/发行方/品牌方", "目标市场/渠道", "动态类型", "动态日期",
    "动态摘要", "建议关注点", "信源等级", "来源层级", "主信源名称", "主信源链接",
    "链接状态", "是否建议入库", "复核备注", "信源数",
]

CUSTOMERS = [
    ["C001", WEEK, SEARCH_DATE, "EC", "Consumer Tech", "电视与显示", "MiniLED电视/世界杯场景",
     "Hisense RGB MiniLED TV pop-up", "Hisense", "美国纽约Hudson Yards/美国官网",
     "线下活动/体育IP营销", "2026-06-12", "Hisense纽约世界杯主题快闪展示RGB MiniLED",
     "世界杯场景、线下互动、美国零售转化", "A", "P0", "PRNewswire",
     "https://www.prnewswire.com/news-releases/hisense-rgb-themed-pop-up-event-lights-up-the-new-landmark-of-fifa-world-cup-2026-302798850.html",
     "可访问", "是", "官方赞助权益与零售导流明确", 1],
    ["C002", WEEK, SEARCH_DATE, "EC", "Consumer Tech", "投影与家庭影院", "4K激光投影",
     "JMGO N3 Ultimate", "JMGO", "全球官网/美国Amazon/足球观赛场景",
     "全球发布/场景营销", "2026-06-11", "JMGO以比赛日场景推广N3 Ultimate三色激光投影",
     "体育观赛、家庭影院、测评内容", "A", "P0", "PRNewswire",
     "https://www.prnewswire.com/news-releases/upgrade-your-match-day-setup-with-jmgo-triple-laser-projectors-302796832.html",
     "可访问", "是", "中国投影品牌；全球发布与观赛场景清晰", 1],
    ["C003", WEEK, SEARCH_DATE, "EC", "FMCG", "食品与饮料", "茶饮/礼盒",
     "Dofo Tea refreshed packaging", "Chaozhou DoFo Tea", "美国官网/海外消费者",
     "品牌焕新/包装升级", "2026-06-12", "Dofo Tea完成全线包装焕新并面向美国订单促销",
     "中国茶文化出海、礼赠包装、DTC转化", "A", "P0", "PRNewswire",
     "https://www.prnewswire.com/news-releases/dofo-tea-refreshes-its-look-to-make-a-centuries-old-tea-ritual-feel-at-home-in-any-kitchen-302796041.html",
     "可访问", "是", "美国免邮与官网转化信息明确", 1],
    ["C004", WEEK, SEARCH_DATE, "EC", "Fashion", "跨境电商平台", "卖家合规/服饰平台",
     "SHEIN Product Safety and Quality Compliance Series", "SHEIN", "全球Marketplace卖家",
     "卖家教育/合规建设", "2026-06-10", "SHEIN联合TIC机构推出卖家产品安全与质量合规培训",
     "平台招商、卖家运营、合规内容", "A", "P0", "PRNewswire",
     "https://www.prnewswire.com/news-releases/shein-expands-marketplace-seller-education-offering-with-product-safety-and-quality-compliance-series-302796523.html",
     "可访问", "是", "Fashion已按独立一级行业收纳", 1],
    ["C005", WEEK, SEARCH_DATE, "EC", "Fashion", "服饰零售", "校园大使/美国内容营销",
     "SHEIN Campus Retreat 2026", "SHEIN", "美国Miami/校园大使",
     "品牌活动/达人社群", "2026-06-04", "SHEIN在迈阿密举办Campus Retreat与首个校园时装秀",
     "美国校园内容、KOC社群、服饰新品曝光", "A", "P0", "PRNewswire",
     "https://www.prnewswire.com/news-releases/shein-returns-to-miami-for-2026-campus-retreat-featuring-first-ever-campus-fashion-show-302791094.html",
     "可访问", "是", "海外活动与目标人群明确", 1],
    ["C006", WEEK, SEARCH_DATE, "AG", "泛娱乐", "长视频/流媒体", "AI影视制作",
     "None Shall Escape / NadouPro", "iQIYI", "iQIYI International",
     "内容上线/AI生产工具", "2026-06-09", "iQIYI上线AI辅助制作电影并同步国际站",
     "国际站内容供给、AI影视生产效率", "A", "P0", "PRNewswire",
     "https://www.prnewswire.com/news-releases/50-more-efficient-iqiyi-emerging-film-project-none-shall-escape-charts-a-new-path-for-ai-film-production-with-nadoupro-302798879.html",
     "可访问", "是", "国际站上线与AI工具事实明确", 1],
    ["C007", WEEK, SEARCH_DATE, "AG", "Gaming", "游戏", "发行/发布会",
     "Tencent Games SPARK 2026 updates", "Tencent Games", "中国及全球市场",
     "发布会/产品更新", "2026-05-29", "Tencent Games在SPARK 2026发布45项国内外市场更新",
     "全球发行储备、IP合作、游戏增长节点", "A", "P0", "Tencent",
     "https://www.tencent.com/en-us/articles/2202340.html",
     "可访问", "是", "官方页面；作为W25游戏线索池前置信号", 1],
    ["C008", WEEK, SEARCH_DATE, "AG", "Gaming", "游戏", "PC/主机叙事动作",
     "Blood Message", "NetEase Games / 24 Entertainment", "全球PC/主机玩家",
     "预告发布/展会曝光", "2026-06-05", "NetEase在Summer Game Fest发布Blood Message新预告",
     "主机化叙事、海外媒体曝光、愿望单承接", "A", "P0", "NetEase Games",
     "https://www.neteasegames.com/news/20260608/37000_1303293.html",
     "可访问", "是", "官方新闻；海外展会曝光明确", 1],
    ["C009", WEEK, SEARCH_DATE, "AG", "Fintech", "跨境支付", "钱包网络/拉美商户",
     "Alipay+ Latin America payments", "Ant International", "智利/阿根廷/拉美商户",
     "区域扩张/支付合作", "2026-05-26", "Alipay+支持合作钱包在拉美PVS商户扫码支付",
     "拉美商户覆盖、旅行支付、钱包伙伴", "A", "P0", "Business Wire",
     "https://www.businesswire.com/news/home/20260526146133/en/Ant-Internationals-Alipay-Enables-Mobile-Payments-for-Global-Travellers-in-Latin-America",
     "可访问", "是", "近14天窗口内的硬信源，可延续到W25跟进", 1],
    ["C010", WEEK, SEARCH_DATE, "AG", "Fintech", "跨境汇款", "汇款到中国",
     "TenPay Global Remit to China", "Tencent Financial Technology / TenPay Global", "新加坡/非中国公民",
     "新服务上线", "2026-06-05", "TenPay Global推出非中国公民汇款到中国服务",
     "跨境汇款、在华收款、金融合规", "B", "P1", "The Asian Banker",
     "https://www.theasianbanker.com/mediafeed-news/details?filter=23792&pd=05+Jun+2026&rkey=20260604AE76132",
     "可访问", "是", "需补PRNewswire原始链接", 1],
    ["C011", WEEK, SEARCH_DATE, "AG", "AI应用", "AI视频生成", "日本AI动画创作者",
     "Hailuo 2.3 / MiniMax", "MiniMax", "日本/全球开发者与创作者",
     "会议披露/产品进展", "2026-05-21", "MiniMax披露Hailuo 2.3在日本AI动画创作者中渗透",
     "日本创作者生态、AI视频订阅、模型API", "A", "P0", "MiniMax",
     "https://www.minimax.io/news/minimax-at-the-waytoagi-global-ai-conference",
     "可访问", "是", "官方博客；时间略早但可做W25预备追踪", 1],
    ["C012", WEEK, SEARCH_DATE, "EC", "Consumer Tech", "户外电源", "便携储能/EOFY促销",
     "BLUETTI EOFY Sale", "BLUETTI", "澳大利亚DTC/线上渠道",
     "区域促销/零售活动", "2026-06-01", "BLUETTI澳洲EOFY促销覆盖Elite 300等储能产品",
     "澳洲大促、便携储能、DTC活动", "A", "P0", "PRNewswire APAC",
     "https://en.prnasia.com/releases/apac/bluetti-launches-eofy-sale-on-power-solutions-with-up-to-44-off-535170.shtml",
     "可访问", "是", "W24已跟踪FridgePower，W25延伸到澳洲促销窗口", 1],
]

SOURCE_HEADERS = [
    "信源ID", "候选ID", "周次", "动态日期", "发布日期/事件日期", "来源层级",
    "信源等级", "信源名称", "URL", "证据摘要", "链接状态/复核备注",
]

SOURCES = [
    [f"S{i:03d}", row[0], WEEK, row[11], row[11], row[15], row[14], row[16], row[17], row[12], row[18]]
    for i, row in enumerate(CUSTOMERS, start=1)
]

EVENT_HEADERS = [
    "展会ID", "周次", "展会时间窗", "行业", "展会/会议", "地点", "日期",
    "展会窗口价值", "信源等级", "链接", "复核备注",
]

EVENTS = [
    ["E001", WEEK, "2026-W25", "跨境电商", "2026 CCEF中国广州跨境电商交易会", "广州", "2026-06-16",
     "跨境卖家、平台服务商、品牌出海客户集中。", "B", "https://www.huodongxing.com/event/9852250280500", "W24沿用至W25窗口"],
    ["E002", WEEK, "2026-W25", "跨境电商", "TikTok Shop ACE全域增长峰会深圳站", "深圳", "2026-06-17",
     "TikTok Shop商家和服务商集中，可抓直播与内容增长线索。", "B", "https://www.amz123.com/hd/jlLZxvTP", "W25窗口"],
    ["E003", WEEK, "2026-W25", "跨境电商", "2026宁波出口跨境电商博览会", "宁波", "2026-06-17",
     "出口跨境电商品牌和服务商集中，可抓展商名单。", "B", "https://www.amz123.com/hd/gjLixPhZ", "W25窗口"],
    ["E004", WEEK, "2026-W25", "独立站", "出海必备：独立站Shopify前沿机会", "杭州", "2026-06-18",
     "独立站卖家和Shopify服务商集中，可抓建站与转化线索。", "B", "https://hz.huodongxing.com/event/8856869193022", "W25窗口"],
]

BID_HEADERS = ["序号", "项目名", "发布方", "业务范畴", "周次", "投标周期", "预算规模", "原始链接"]

BIDS = [
    [1, "九寨沟国际传播中心海外宣传营销项目", "九寨沟管理局", "海外宣传营销/国际传播", WEEK,
     "2026-06-05至2026-06-17", "没披露",
     "https://www.jiuzhai.com/dynamic/public/11162-2026-06-05-08-33-03"],
    [2, "泸州老窖·国窖1573海外品牌推广咨询区域性服务项目", "泸州老窖", "海外品牌推广咨询", WEEK,
     "没披露", "没披露", "https://www.lzlj.com/news/bulletin/tenders/12576.html"],
]

EXCLUDED_HEADERS = ["排除ID", "一级行业", "主体/项目", "动态类型", "排除原因"]
EXCLUDED = [
    ["X001", "智能出行", "AION UK launch", "整车出海", "车企整车发布不进入当前电商商品或应用候选主表。"],
    ["X002", "韩国游戏", "Ragnarok: Twilight Global", "游戏OBT", "非中国出海主体，先作为Gaming外围观察，不进入主表。"],
]

SEARCH_HEADERS = ["检索ID", "任务", "主要信源域", "代表性检索式/动作", "检索日期", "结果摘要"]
SEARCH_ROWS = [
    ["R001", "W25预备雷达", "PRNewswire / BusinessWire / 官方新闻", "China brand global launch June 2026", SEARCH_DATE, "客户候选12条"],
    ["R002", "Gaming雷达", "Tencent / NetEase / PRNewswire", "Tencent SPARK NetEase Summer Game Fest", SEARCH_DATE, "Gaming候选2条"],
    ["R003", "展会窗口", "W24展会线索延续", "2026-06-16至2026-06-22", SEARCH_DATE, "展会4条"],
    ["R004", "排除纠偏", "PRNewswire / 搜索结果", "整车/非中国主体过滤", SEARCH_DATE, "排除2条"],
]

RULE_ROWS = [
    ["文件名", OUT.name],
    ["周次口径", "W25文件承载W25预备客户动态、W25展会窗口及近期招投标线索"],
    ["客户动态", "仅收中国出海应用/商品/品牌或明确海外渠道动作"],
    ["网页消费", "门户只读取portal/data/weekly下JSON，不直接读取Excel"],
    ["规则源", r"Z:\增长分析中台\线索采集\客户动态检索规则_最新版.md"],
]


def add_sheet(wb, title, headers, rows):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for row in rows:
        ws.append(row)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
    ws.freeze_panes = "A2"
    for idx, header in enumerate(headers, start=1):
        max_len = max([len(str(header))] + [len(str(row[idx - 1])) if idx - 1 < len(row) else 0 for row in rows])
        ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 10), 42)


def main():
    wb = Workbook()
    wb.remove(wb.active)
    add_sheet(wb, "客户候选名单", CUSTOMER_HEADERS, CUSTOMERS)
    add_sheet(wb, "信源明细", SOURCE_HEADERS, SOURCES)
    add_sheet(wb, "展会线索", EVENT_HEADERS, EVENTS)
    add_sheet(wb, "招投标线索", BID_HEADERS, BIDS)
    add_sheet(wb, "排除观察", EXCLUDED_HEADERS, EXCLUDED)
    add_sheet(wb, "检索记录", SEARCH_HEADERS, SEARCH_ROWS)
    add_sheet(wb, "规则版本", ["字段", "内容"], RULE_ROWS)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(OUT)


if __name__ == "__main__":
    main()
