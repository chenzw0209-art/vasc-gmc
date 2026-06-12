import json
from collections import Counter
from datetime import date, datetime
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "data_assets" / "canonical_sources" / "gaming" / "raw_2026_06_09" / "calendar" / "新游日历目标游戏枚举_v0.2_20260608.xlsx"
OUT = ROOT / "portal" / "data" / "gaming" / "new_game_calendar_2026_06_08.json"
BLUEPRINT_OUT = ROOT / "portal" / "data" / "gaming" / "gaming_calendar_targets_2026_06_09.json"


def clean(value):
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()[:10]
    return str(value).strip()


def sheet_rows(workbook, name):
    ws = workbook[name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [clean(x) for x in rows[0]]
    records = []
    for raw in rows[1:]:
        item = {headers[i]: clean(raw[i]) for i in range(min(len(headers), len(raw))) if headers[i]}
        if any(item.values()):
            records.append(item)
    return records


def count(records, field):
    return dict(Counter(clean(r.get(field)) or "待补充" for r in records))


def split_platform(platform):
    return [x.strip() for x in clean(platform).replace("/", ",").split(",") if x.strip()]


def priority_rank(value):
    return {"P0": 0, "P1": 1, "P2": 2}.get(clean(value), 9)


def confidence_rank(value):
    return {"高": 0, "中高": 1, "中": 2, "低": 3}.get(clean(value), 9)


def build_action_copy(item):
    reason = item.get("bd_reason") or item.get("attention_point") or item.get("dynamic_summary")
    window = item.get("bd_window") or item.get("estimated_window") or "待补充窗口"
    if not reason:
        reason = "需补充 BD 切入理由"
    return f"{item.get('game_name', '该项目')}处于{item.get('current_stage') or item.get('stage') or '待补充阶段'}，窗口为{window}。建议从{reason}切入。"


def source_refs_for_item(item, source_map):
    refs = []
    lead_sources = source_map.get(item.get("lead_id"), [])
    for source in lead_sources[:4]:
        refs.append({
            "source_id": source.get("source_id", ""),
            "source_type": source.get("source_type", ""),
            "source_grade": source.get("source_grade", ""),
            "source_name": source.get("source_name", ""),
            "url": source.get("url", ""),
            "evidence_summary": source.get("evidence_summary", ""),
            "dynamic_date": source.get("dynamic_date", ""),
            "is_primary": source.get("is_primary", ""),
        })
    if not refs and item.get("primary_source_url"):
        refs.append({
            "source_id": f"{item.get('lead_id', '')}_primary",
            "source_type": "主信源",
            "source_grade": item.get("source_grade", ""),
            "source_name": "主信源",
            "url": item.get("primary_source_url", ""),
            "evidence_summary": item.get("evidence_summary", ""),
            "dynamic_date": item.get("recent_event_date", ""),
            "is_primary": "是",
        })
    return refs


def item_card(item, source_map):
    refs = source_refs_for_item(item, source_map)
    return {
        "id": item.get("lead_id") or item.get("calendar_id"),
        "calendar_id": item.get("calendar_id", ""),
        "game_name": item.get("game_name", ""),
        "publisher": item.get("owner", ""),
        "category": item.get("category", ""),
        "platforms": item.get("platforms", []),
        "calendar_bucket": item.get("calendar_bucket", ""),
        "stage": item.get("stage", ""),
        "current_stage": item.get("current_stage", ""),
        "playable_level": item.get("playable_level", ""),
        "estimated_date": item.get("estimated_date", ""),
        "estimated_window": item.get("estimated_window", ""),
        "bd_priority": item.get("bd_priority", ""),
        "bd_window": item.get("bd_window", ""),
        "bd_reason": item.get("bd_reason", ""),
        "action_copy": build_action_copy(item),
        "dynamic_summary": item.get("dynamic_summary", ""),
        "attention_point": item.get("attention_point", ""),
        "detail_summary": item.get("detail_summary", ""),
        "evidence_summary": item.get("evidence_summary", ""),
        "source_grade": item.get("source_grade", ""),
        "confidence": item.get("confidence", ""),
        "layer": item.get("layer", ""),
        "evidence_refs": refs,
        "source_drawer": {
            "origin_type": "governed",
            "source_types": sorted({x.get("source_type", "") for x in refs if x.get("source_type")}),
            "source_count": len(refs),
            "confidence": item.get("confidence", ""),
            "verification_metrics": [
                "官方阶段或商店页更新",
                "后续周度投放/素材是否升温",
                "上线窗口是否被公开信源确认",
            ],
            "downgrade_signal": "公开节点延后、主信源失效，或后续周度投放/下载/收入均无跟进。",
        },
    }


def build_blueprint_output(summary, calendar_items, source_map, stage_rows):
    cards = [item_card(item, source_map) for item in calendar_items]
    cards.sort(key=lambda x: (
        priority_rank(x.get("bd_priority")),
        x.get("estimated_date") or "9999-12-31",
        confidence_rank(x.get("confidence")),
        x.get("game_name"),
    ))

    dated_items = [x for x in cards if x.get("estimated_date")]
    window_pool = [x for x in cards if not x.get("estimated_date")]
    calendar_days = {}
    for item in dated_items:
        calendar_days.setdefault(item["estimated_date"], []).append({
            "id": item["id"],
            "game_name": item["game_name"],
            "bd_priority": item["bd_priority"],
            "calendar_bucket": item["calendar_bucket"],
            "stage": item["stage"],
        })

    focus_cards = [x for x in cards if x.get("bd_priority") in {"P0", "P1"}]
    p0_cards = [x for x in cards if x.get("bd_priority") == "P0"]

    return {
        "version": "v0.1",
        "industry": "Gaming",
        "module": "new_game_calendar",
        "period": "2026-06-09",
        "generated_at": "2026-06-09",
        "schema_role": "portal_consumption",
        "source_registry": [
            {
                "source_id": "G-CAL-001",
                "source_type": "new_game_calendar",
                "source_path": str(SOURCE.relative_to(ROOT)),
                "role": "future_opportunity_and_bd_window",
            }
        ],
        "governance": {
            "frontend_reads_json_only": True,
            "no_frontend_category_derivation": True,
            "origin_type": "governed",
            "source_excel_is_offline_only": True,
        },
        "coverage": {
            "calendar_items": summary["calendar_items"],
            "effective_targets": summary["effective_targets"],
            "watchlist": summary["watchlist"],
            "source_count": summary["source_count"],
            "known_limits": [
                "新游日历只能识别机会窗口，不能单独证明收入规模。",
                "预计日期是治理层点估或公开窗口，不等同于确定上线日期。",
                "BD 切入理由来自线索/行研治理字段，需要结合后续公开信号复核。",
            ],
        },
        "modules": {
            "core_thesis": {
                "title": "新游窗口正在前置 Gaming 出海的 BD 机会",
                "body": "本模块用于识别处于预约、测试、Demo、愿望单、EA 或上线前预热阶段的项目；它回答谁进入行动窗口，不直接证明谁已经商业化成功。",
                "verification_metrics": [
                    "P0/P1 项目是否出现公开阶段证据",
                    "后续 Insight 周度投放或素材是否升温",
                    "上线窗口是否被官方或商店页持续确认",
                ],
                "downgrade_signal": "公开节点延后、证据链接失效，或后续周度数据无任何投放/下载/收入跟进。",
            },
            "calendar": {
                "display_mode": "calendar",
                "calendar_days": calendar_days,
                "window_pool": window_pool,
                "tooltip_fields": [
                    "game_name",
                    "publisher",
                    "current_stage",
                    "estimated_window",
                    "bd_priority",
                    "bd_window",
                    "bd_reason",
                    "evidence_summary",
                ],
            },
            "focus_cards": focus_cards,
            "p0_cards": p0_cards,
            "source_drawer_index": {
                card["id"]: card["source_drawer"] | {"evidence_refs": card["evidence_refs"]}
                for card in cards
            },
            "summary_cards": [
                {"label": "日历项目", "value": summary["calendar_items"], "note": "进入新游日历视图"},
                {"label": "P0 项目", "value": summary["p0_items"], "note": "优先建联窗口"},
                {"label": "有效目标", "value": summary["effective_targets"], "note": "已过基础证据门槛"},
                {"label": "观察池", "value": summary["watchlist"], "note": "待补证项目"},
            ],
            "counts": {
                "calendar_bucket_counts": summary["calendar_bucket_counts"],
                "priority_counts": summary["priority_counts"],
                "stage_counts": summary["stage_counts"],
                "platform_counts": summary["platform_counts"],
            },
            "stage_dictionary": stage_rows,
        },
        "copy_templates": {
            "calendar_tooltip": "该项目处于【当前阶段】，预计窗口为【预计上线窗口】。当前 BD 优先级为【BD优先级】，建议从【BD切入理由】切入。",
            "risk_note": "如果公开节点延后且后续周度投放、下载或收入均无跟进，则该项目从行动窗口降级为观察池。",
        },
    }


def main():
    if not SOURCE.exists():
        raise FileNotFoundError(f"Missing Gaming calendar source: {SOURCE}")
    wb = openpyxl.load_workbook(SOURCE, read_only=True, data_only=True)
    dashboard = list(wb["00_看板"].iter_rows(values_only=True))
    calendar_rows = sheet_rows(wb, "06_日历视图")
    target_rows = sheet_rows(wb, "02_有效目标39")
    watch_rows = sheet_rows(wb, "03_观察待补证23")
    source_rows = sheet_rows(wb, "05_信源明细")
    stage_rows = sheet_rows(wb, "07_阶段字典")

    target_by_id = {r.get("lead_id"): r for r in target_rows}
    watch_by_id = {r.get("lead_id"): r for r in watch_rows}

    calendar_items = []
    for r in calendar_rows:
        lead_id = r.get("lead_id", "")
        detail = target_by_id.get(lead_id) or watch_by_id.get(lead_id) or {}
        calendar_items.append({
            "calendar_id": r.get("日历ID", ""),
            "lead_id": lead_id,
            "game_name": r.get("游戏名", ""),
            "owner": r.get("归属方") or detail.get("归属方_展示") or detail.get("发行/母公司", ""),
            "platform": r.get("平台", ""),
            "platforms": split_platform(r.get("平台", "")),
            "calendar_bucket": r.get("日历视图归属", ""),
            "stage": r.get("标准阶段", ""),
            "playable_level": r.get("实机等级", ""),
            "estimated_date": r.get("预计日期_点估", ""),
            "estimated_window": r.get("预计窗口", ""),
            "confidence": r.get("置信度", ""),
            "bd_priority": r.get("BD优先级", ""),
            "recent_event_date": r.get("最近事件日期", ""),
            "dynamic_summary": r.get("动态摘要_短", ""),
            "primary_source_url": r.get("主信源URL", ""),
            "rule_note": r.get("规则备注", ""),
            "category": detail.get("品类", ""),
            "current_stage": detail.get("当前阶段", ""),
            "bd_window": detail.get("拓客窗口", ""),
            "bd_reason": detail.get("BD切入理由", ""),
            "attention_point": detail.get("建议关注点_短", ""),
            "detail_summary": detail.get("详情摘要", ""),
            "evidence_summary": detail.get("证据摘要_详情", ""),
            "source_grade": detail.get("信源等级") or "",
            "layer": detail.get("新规则处理层级") or "",
        })

    source_map = {}
    for r in source_rows:
        source_map.setdefault(r.get("lead_id", ""), []).append({
            "source_id": r.get("信源ID", ""),
            "source_type": r.get("证据类型", ""),
            "source_grade": r.get("信源等级", ""),
            "source_name": r.get("主信源名称", ""),
            "url": r.get("URL", ""),
            "evidence_summary": r.get("证据摘要", ""),
            "dynamic_date": r.get("动态日期", ""),
            "is_primary": r.get("是否主证据", ""),
        })

    p0 = sum(1 for r in calendar_items if r["bd_priority"] == "P0")
    buckets = count(calendar_items, "calendar_bucket")
    priorities = count(calendar_items, "bd_priority")
    stages = count(calendar_items, "stage")
    platforms = Counter()
    for r in calendar_items:
        for p in r["platforms"]:
            platforms[p] += 1

    summary = {
        "generated_at": "2026-06-08",
        "source_file": str(SOURCE),
        "total_pool": clean(dashboard[2][1]) if len(dashboard) > 2 else "",
        "effective_targets": len(target_rows),
        "watchlist": len(watch_rows),
        "calendar_items": len(calendar_items),
        "source_count": len(source_rows),
        "p0_items": p0,
        "high_confidence_items": sum(1 for r in calendar_items if r["confidence"] in {"高", "中高"}),
        "a_grade_sources": sum(1 for r in target_rows + watch_rows if r.get("信源等级") == "A"),
        "calendar_bucket_counts": buckets,
        "priority_counts": priorities,
        "stage_counts": stages,
        "platform_counts": dict(platforms.most_common(8)),
    }

    output = {
        "version": "v0.1",
        "industry": "Gaming",
        "module": "新游日历",
        "summary": summary,
        "calendar_items": calendar_items,
        "target_games": target_rows,
        "watchlist_games": watch_rows,
        "source_map": source_map,
        "stage_dictionary": stage_rows,
    }

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(output, ensure_ascii=False, indent=2), encoding="utf-8")
    blueprint_output = build_blueprint_output(summary, calendar_items, source_map, stage_rows)
    BLUEPRINT_OUT.write_text(json.dumps(blueprint_output, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {OUT} ({len(calendar_items)} calendar items)")
    print(f"Wrote {BLUEPRINT_OUT} ({len(blueprint_output['modules']['focus_cards'])} focus cards)")


if __name__ == "__main__":
    main()
