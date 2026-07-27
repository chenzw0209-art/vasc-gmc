#!/usr/bin/env python3
"""Build weekly portal JSON files from the lead-collection workbook."""

from __future__ import annotations

import argparse
import json
import re
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel


CUSTOMER_MAP = {
    "候选ID": "candidate_id",
    "周次": "week",
    "检索日期": "search_date",
    "钛动分类": "titan_category",
    "一级行业": "standard_l1",
    "二级行业": "standard_l2",
    "三级品类": "standard_l3",
    "应用/商品名称": "application_product_name",
    "背后公司/发行方/品牌方": "owner_company_brand",
    "目标市场/渠道": "target_market_channel",
    "动态类型": "dynamic_type",
    "动态日期": "dynamic_date",
    "动态摘要": "dynamic_summary",
    "建议关注点": "attention_point",
    "信源等级": "source_grade",
    "来源层级": "source_layer",
    "主信源名称": "primary_source_name",
    "主信源链接": "primary_source_url",
    "链接状态": "link_status",
    "是否建议入库": "warehouse_suggestion",
    "复核备注": "review_note",
    "信源数": "source_count",
}

SOURCE_MAP = {
    "信源ID": "source_id",
    "候选ID": "candidate_id",
    "周次": "week",
    "动态日期": "dynamic_date",
    "发布日期/事件日期": "publish_event_date",
    "来源层级": "source_layer",
    "信源等级": "source_grade",
    "信源名称": "source_name",
    "URL": "url",
    "证据摘要": "evidence_summary",
    "链接状态/复核备注": "link_status_review_note",
}

EVENT_MAP = {
    "展会ID": "event_id",
    "周次": "week",
    "展会时间窗": "event_window",
    "行业": "industry",
    "展会/会议": "event_name",
    "地点": "location",
    "日期": "date",
    "信源等级": "source_grade",
    "链接": "url",
    "参加这个展会能获得什么": "event_value",
    "复核备注": "review_note",
}

TENDER_MAP = {
    "序号": "tender_id",
    "项目名": "project_name",
    "发布方": "publisher",
    "业务范畴": "business_scope",
    "周次": "week",
    "投标周期": "bid_period",
    "预算规模": "budget",
    "原始链接": "url",
}

WEEKLY_SUMMARY_OVERRIDES = {
    "W31": "\n".join([
        "Tencent / Weixin：面向东南亚启动小程序全球创新挑战赛区域赛，并引入AI工具支持青年开发者；建议 EC、AI应用与本地化团队关注区域生态拓展、开发者运营和品牌合作机会。",
        "DJI / 大疆：以机场设备支持澳大利亚矿业与农业的大范围超视距无人机部署；建议 Consumer Tech 与区域业务团队关注澳洲行业场景规模化、本地伙伴和内容传播需求。",
        "2026中国国际数码互动娱乐展览会（ChinaJoy）：7月31日至8月3日在上海举行，游戏、内容平台与品牌资源集中；建议 AG、Gaming 与内容营销团队重点获取发行商、平台方和全球化合作线索。",
    ]),
    "W30": "\n".join([
        "华为：7月14日在吉隆坡举行全球旗舰新品发布会，Pura 90s 系列与穿戴、平板同步亮相；新品上市期适合围绕影像内容、创作者测评与区域社媒节奏切入。",
        "2026悉尼Online Retailer Exhibition：7月22日至23日在 ICC Sydney 举办，聚集电商、零售、营销与数据服务主体；适合优先获取澳洲电商生态中的展商联系人与合作需求。",
        "杭州市文化广电旅游局：发布海外传播矩阵推广采购，直接覆盖国际内容传播与海外推广；与达人内容、区域化创意和投放协同能力高度匹配。",
    ]),
    "W29": "\n".join([
        "2026 世界人工智能大会（7月17-20日，上海）：展会进入近场窗口，已从展商名录复核出48家具有红人营销适配度的目标客户；应优先安排现场触达和会后跟进。",
        "江苏徐工电子商务股份有限公司：启动TikTok运营服务招标，覆盖3个账号、内容运营、广告投放和海外询盘转化，最高限价36万元；这是工程机械出海效果营销的直接需求。",
        "中免集团（海南）运营总部有限公司：公开招标小红书广告投放及KOS投流，服务期1年、7月28日截标；项目与达人/KOS内容投放能力高度匹配，值得优先评估投标。",
    ]),
    "W27": "\n".join([
        "Huawei：与泰国7-Eleven运营方CP All签署合作，将智能穿戴支付接入泰国零售门店并面向中国游客场景；这是支付、零售与旅游营销预算同时可触达的东南亚渠道信号。",
        "Segway Commercial：联合Whoosh在墨西哥世界杯期间推出免费骑行活动，把共享出行设备与全球赛事流量绑定；适合跟进拉美本地运营、体育营销和城市出行合作需求。",
        "CStone Pharmaceuticals：与Arrotex签署Sugemalimab在澳大利亚和新西兰的独家商业化协议；中国创新药出海已进入市场准入和本地商业化执行阶段，值得优先跟进医药国际化预算。",
    ]),
    "W26": "\n".join([
        "OPPO：Reno16 Series启动全球发布，并宣布BABYMONSTER担任Reno全球品牌大使；印度与东南亚市场的新品承接，是本周最明确的手机出海信号。",
        "MiniMax：推出面向全球创作者的一站式AI视频生成工具Hub；生成式视频产品正从模型能力展示，转向可直接使用的创作工作台。",
        "上海世界移动通信大会：6月24日在上海新国际博览中心开幕；通信、AI与智能硬件品牌集中，是本周获取展商与新品发布线索最密集的线下窗口。",
    ]),
    "W25": "\n".join([
        "客户关注草本健康保健赛道出海，HEC Life+、东阳光（HEC Life+），推测扩张SEA。",
        "展会关注2026 CCEF中国(广州)跨境电商交易会（广州跨交会）。",
        "招投标关注ToG，营销方面季度中频繁释放出出海信号，如泸州老窖、广州酒家。",
        "Gaming新客日历枚举版本上线（详见Gaming）。",
    ]),
}


def cell_value(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if value is None:
        return ""
    return str(value).strip()


def read_sheet(wb, sheet_name, mapping):
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    headers = [cell_value(c.value) for c in ws[1]]
    rows = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        raw = {headers[i]: cell_value(values[i]) for i in range(min(len(headers), len(values)))}
        if not any(raw.values()):
            continue
        row = {target: raw.get(source, "") for source, target in mapping.items()}
        rows.append(row)
    return rows


def mmdd(value):
    text = str(value or "")
    try:
        return datetime.fromisoformat(text).strftime("%m-%d")
    except ValueError:
        return text[:5] if text else ""


def parse_sheet_date(value):
    text = str(value or "").strip()
    if not text:
        return None
    try:
        return datetime.fromisoformat(text).date()
    except ValueError:
        pass
    if re.fullmatch(r"\d+(\.\d+)?", text):
        try:
            return from_excel(float(text)).date()
        except (TypeError, ValueError):
            return None
    return None


def parse_iso_dates(text):
    dates = []
    for match in re.findall(r"20\d{2}-\d{2}-\d{2}", str(text or "")):
        try:
            dates.append(datetime.fromisoformat(match).date())
        except ValueError:
            pass
    return dates


def filter_active_tenders(tenders, generated_at):
    try:
        today = datetime.fromisoformat(generated_at).date()
    except ValueError:
        today = date.today()

    closed_keywords = ("成交公告", "中标结果", "中标候选", "结果公告")
    stale_month_markers = ("2026-05",)
    active = []
    for row in tenders:
        text = " ".join([
            row.get("project_name", ""),
            row.get("publisher", ""),
            row.get("business_scope", ""),
            row.get("bid_period", ""),
        ])
        if any(keyword in text for keyword in closed_keywords):
            continue
        if any(marker in row.get("bid_period", "") for marker in stale_month_markers):
            continue
        dates = parse_iso_dates(row.get("bid_period", ""))
        if dates and max(dates) < today:
            continue
        active.append(row)
    return active


VALID_EC_L2 = {
    "Beauty": {
        "面部护理", "身体与防晒护理", "头发护理", "彩妆", "口腔护理", "香水", "剃须脱毛产品",
        "指甲与足部护理", "美容工具与配件", "除臭与个护小品",
    },
    "Consumer Tech": {
        "手机与配件", "笔记本电脑", "计算机配件与外设", "笔电与平板配件", "显示器与电脑平板",
        "数据存储与外组件", "计算机网络", "台式与电脑打印机", "办公电子产品", "影像设备",
        "监控与家庭安防摄像头", "穿戴式科技", "耳机", "家庭音频与音响", "电视与视频显示",
        "录音与便携影音", "厨房小家电", "咖啡茶饮电器", "暖通电器", "清洁电器", "美发造型电器",
        "个护医疗电器", "大家电与配件", "电动工具", "照明产品", "安防硬件", "户外便携电源储能",
        "视频游戏机与配件", "汽车电子",
    },
    "FMCG": {
        "饮料", "零食与糖果", "厨房储备食品", "鲜食与日配", "餐饮礼品与酒精", "家用清洁剂",
        "纸品与一次性用品", "洗衣用品", "洗碗用品", "家用清洁工具", "婴儿耗材", "家用电池",
    },
    "Health": {
        "维生素补充剂基础", "草药植物补充剂", "鱼油抗氧化胶原", "肠胃消化睡眠酶", "运动营养",
        "减肥与代餐", "非处方药与急救", "失禁与女性保健", "替代医学与睡眠辅助", "糖尿病与呼吸辅助",
        "家用医疗设备", "视觉护理",
    },
    "Lifestyle": {
        "宠物用品", "家庭存储与组织", "运动服装与设备", "健身设备", "玩具", "家具", "户外娱乐",
        "床上用品", "厨房非电器", "儿童家居与婴幼装备", "艺术绘画手工", "汽车零件配件", "家居装饰",
        "户外家具与装饰", "农场牧场与户外动力", "乐器", "办公学校用品", "户外园艺", "汽车工具与保养",
    },
    "Fashion": {"女装", "男装", "童装", "鞋履", "箱包", "配饰/珠宝/手表"},
}

EC_L2_NORMALIZATION = {
    ("Beauty", "彩妆"): "彩妆",
    ("Consumer Tech", "电视与显示"): "电视与视频显示",
    ("Consumer Tech", "投影与家庭影院"): "电视与视频显示",
    ("Consumer Tech", "户外电源"): "户外便携电源储能",
    ("Consumer Tech", "充电与能源设备"): "汽车电子",
    ("Consumer Tech", "电脑外设与充电"): "计算机配件与外设",
    ("Consumer Tech", "厨房与大家电"): "大家电与配件",
    ("Consumer Tech", "智能车载"): "汽车电子",
    ("Consumer Tech", "智能穿戴"): "穿戴式科技",
    ("Consumer Tech", "赛事显示设备"): "电视与视频显示",
    ("Fashion", "跨境电商平台"): "女装",
    ("Fashion", "服饰零售"): "女装",
    ("FMCG", "食品与饮料"): "饮料",
    ("FMCG", "食品原料"): "厨房储备食品",
    ("Health", "营养与保健"): "草药植物补充剂",
    ("Lifestyle", "家纺寝具"): "床上用品",
    ("Lifestyle", "智能居住"): "家具",
}

INDUSTRY_OVERRIDES = {
    ("EC", "Health", "智能健身"): ("EC", "Lifestyle", "健身设备"),
    ("EC", "工业", "清洁能源"): ("AG", "工业", "-"),
    ("EC", "智能出行", "整车出海"): ("AG", "智能出行", "-"),
}


def normalize_industry(row):
    raw_category = row.get("titan_category", "")
    raw_l1 = row.get("standard_l1", "")
    raw_l2 = row.get("standard_l2", "")
    row["source_titan_category"] = raw_category
    row["source_standard_l1"] = raw_l1
    row["source_standard_l2"] = raw_l2
    override = INDUSTRY_OVERRIDES.get((raw_category, raw_l1, raw_l2))
    if override:
        row["titan_category"], row["standard_l1"], row["standard_l2"] = override
        if row["titan_category"] == "AG":
            row["standard_l2"] = "-"
        return row
    if row.get("titan_category") == "AG":
        row["standard_l2"] = "-"
        return row
    key = (row.get("standard_l1", ""), raw_l2)
    row["standard_l2"] = EC_L2_NORMALIZATION.get(key, raw_l2 or "-")
    return row


def validate_industry_taxonomy(records):
    errors = []
    for row in records:
        candidate_id = row.get("candidate_id", "UNKNOWN")
        category = row.get("titan_category", "")
        l1 = row.get("standard_l1", "")
        l2 = row.get("standard_l2", "")
        if category == "AG":
            if l2 != "-":
                errors.append(f"{candidate_id}: AG standard_l2 must be -, got {l2}")
            continue
        if category != "EC":
            continue
        valid_l2s = VALID_EC_L2.get(l1)
        if not valid_l2s:
            errors.append(f"{candidate_id}: EC standard_l1 not in dictionary: {l1}")
            continue
        if l2 not in valid_l2s:
            source_l1 = row.get("source_standard_l1", l1)
            source_l2 = row.get("source_standard_l2", l2)
            errors.append(f"{candidate_id}: EC {l1}/{l2} not in dictionary; source was {source_l1}/{source_l2}")
    if errors:
        raise ValueError("weekly leads industry taxonomy validation failed:\n- " + "\n- ".join(errors))


def lead_sort_key(row):
    category_order = {"EC": 0, "AG": 1}
    candidate_id = row.get("candidate_id", "")
    return (category_order.get(row.get("titan_category", ""), 9), candidate_id)


def build_weekly(content_rows, source_rows, event_rows, tender_rows, week, generated_at, workbook):
    records = sorted(
        [row for row in content_rows if row.get("warehouse_suggestion", "是") != "否"],
        key=lead_sort_key,
    )
    validate_industry_taxonomy(records)
    for row in records:
        row["dynamic_date_display"] = mmdd(row.get("dynamic_date"))

    by_l3 = defaultdict(list)
    by_l2 = defaultdict(list)
    for row in records:
        by_l3[row.get("standard_l3", "待补")].append(row)
        by_l2[(row.get("standard_l1", "待补"), row.get("standard_l2", "待补"))].append(row)

    industry_brief_by_industry = {}
    for key, rows in by_l3.items():
        first = rows[0]
        industry_brief_by_industry[key] = {
            "industry": key,
            "lead_count": len(rows),
            "headline": f"{first.get('standard_l2', '行业')}本周新增{len(rows)}条可复核信号",
            "signals": [r.get("dynamic_summary", "") for r in rows[:4]],
            "recommended_focus": [r.get("attention_point", "") for r in rows[:4]],
        }

    similar = {}
    for row in records:
        peers = [
            peer for peer in by_l2[(row.get("standard_l1", ""), row.get("standard_l2", ""))]
            if peer.get("candidate_id") != row.get("candidate_id")
        ][:4]
        similar[row.get("candidate_id", "")] = peers

    evidence = defaultdict(list)
    for source in source_rows:
        evidence[source.get("candidate_id", "")].append(source)

    product_count = sum(1 for row in records if row.get("titan_category") == "EC")
    app_count = len(records) - product_count
    grade_a_count = sum(1 for row in records if row.get("source_grade") == "A")
    by_l1 = defaultdict(list)
    for row in records:
        by_l1[row.get("standard_l1", "待补")].append(row)
    top_l1 = sorted(by_l1.items(), key=lambda item: (-len(item[1]), item[0]))[:5]
    top_summary = "；".join(f"{l1}{len(rows)}条" for l1, rows in top_l1) or "本周情报待补充"

    return {
        "generated_at": generated_at,
        "source_workbook": str(workbook),
        "weekly_module_content": {
            "week": f"2026-{week}",
            "top_summary": WEEKLY_SUMMARY_OVERRIDES.get(
                week,
                f"{week}更新：{top_summary}。本周新增出海媒体补扫，重点补强Beauty、Fashion、FMCG等非3C类目。",
            ),
            "kpis": {
                "focus_customer_count": len(records),
                "new_customer_signal_count": len(records),
                "product_candidate_count": product_count,
                "app_candidate_count": app_count,
                "grade_a_count": grade_a_count,
                "exhibition_count": len(event_rows),
                "tender_count": len(tender_rows),
            },
            "focus_customers": records,
            "industry_brief_by_industry": industry_brief_by_industry,
            "similar_customers_by_candidate": similar,
        },
        "leads_module_content": {
            "week": f"2026-{week}",
            "records": records,
        },
        "exhibition_window_content": event_rows,
        "tender_opportunity_content": tender_rows,
        "evidence_chain_detail_mapping": dict(evidence),
        "data_quality_check_result": {
            "sheet_source": "售前情报库周更新工作簿",
            "customer_rows": len(content_rows),
            "event_rows": len(event_rows),
            "tender_rows": len(tender_rows),
            "source_rows": len(source_rows),
            "customer_rows_note": "仅筛除是否建议入库=否的记录；页面只消费本JSON，不读取Excel。",
        },
    }


def verification_metrics(l1):
    if l1 == "Gaming":
        return ["预约/测试转化", "Steam愿望单", "Discord/社媒增速", "上线地区榜单"]
    if l1 == "Consumer Tech":
        return ["Amazon排名变化", "测评内容声量", "官网活动转化", "零售价格带"]
    if l1 == "Fintech":
        return ["覆盖商户数", "钱包/银行伙伴数", "新增国家", "交易成功率"]
    if l1 == "Fashion":
        return ["卖家参与数", "校园/达人内容量", "商品合规节奏", "美国市场活动转化"]
    return ["官方更新频率", "海外渠道信号", "社媒声量", "转化数据"]


def build_industry_supply(records, week, generated_at, workbook):
    grouped = defaultdict(list)
    for row in records:
        grouped[(row.get("standard_l1", "待补"), row.get("standard_l2", "待补"))].append(row)

    industries = []
    for (l1, l2), rows in sorted(grouped.items()):
        top_owner = rows[0].get("owner_company_brand", "待补")
        signals = [r.get("dynamic_summary", "") for r in rows[:5]]
        industries.append({
            "source_lead_industry_key": f"{l1}|{l2}",
            "primary_industry": l1,
            "secondary_industry": l2,
            "mapped_research_industry": l2,
            "industry_key": f"{l1}|{l2}",
            "metrics": {
                "gmv": "本周线索侧未配置行业规模口径",
                "yoy_growth": "本周线索侧未配置同比口径",
                "cn_gmv_share": "本周线索侧未配置中国品牌份额口径",
                "top_cn_player": top_owner,
            },
            "current_stage": "周度线索扫描",
            "main_variable": f"{l2}本周变量来自{len(rows)}条客户动态，优先用于售前名单筛选和触达优先级判断。",
            "main_tension": "当前供给来自周度客户动态，行业规模和份额不在本周情报口径内。",
            "market_misread": "不要把单条发布误读为行业趋势，先看同类主体是否连续出现。",
            "growth_signals_short": signals,
            "verification_metrics": verification_metrics(l1),
            "counter_signals": ["硬信源后续没有渠道/榜单数据承接", "海外市场动作停留在传播层"],
            "sales_translation": "优先围绕新品、测试、区域发布、线下活动和渠道合作切入，做客户名单补全与复核。",
            "similar_customers": [{"brand": r.get("owner_company_brand", ""), "grade": r.get("source_grade", "")} for r in rows[:4]],
            "this_week_leads": [{
                "candidate_id": r.get("candidate_id", ""),
                "lead_category": r.get("titan_category", ""),
                "primary_industry": r.get("standard_l1", ""),
                "secondary_industry": r.get("standard_l2", ""),
                "tertiary_category": r.get("standard_l3", ""),
                "product_name": r.get("application_product_name", ""),
                "owner": r.get("owner_company_brand", ""),
                "dynamic_type": r.get("dynamic_type", ""),
                "dynamic_summary": r.get("dynamic_summary", ""),
                "suggested_focus": r.get("attention_point", ""),
                "source_grade": r.get("source_grade", ""),
            } for r in rows],
            "research_source_path": f"线索采集正式库 {week} 客户动态与信源明细",
            "bottom_table_path": str(workbook),
            "coverage_status": "周度线索覆盖",
            "industry_conclusion": f"{l1}/{l2}本周有{len(rows)}条可复核客户动态，适合先进入销售扫描池。",
            "visual_summary": {
                "title": f"{l1}/{l2}：{week}周度信号",
                "metric_line": f"{len(rows)}条线索 / Top主体：{top_owner}",
                "signal_line": "；".join(signals[:2]),
                "risk_line": "本页只呈现周度线索事实，不替代行业规模判断。",
            },
            "this_week_lead_count": len(rows),
        })

    return {
        "week": f"2026-{week}",
        "generated_at": generated_at,
        "source_files": {
            "lead_workbook": str(workbook),
            "research_root": "Z:\\增长分析中台\\行研洞察",
        },
        "industries": industries,
        "quality_check": {
            "lead_rows": len(records),
            "industry_group_count": len(industries),
            "covered_industry_count": 0,
            "desktop_research_count": len(industries),
            "gap_industry_count": 0,
            "missing_required_fields": [],
            "notes": [
                f"{week}行业供给由本周客户线索聚合生成，仅用于售前扫描与同类主体归并。",
                "页面只读取portal/data/weekly下JSON，不读取Excel。",
            ],
        },
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", required=True)
    parser.add_argument("--week", required=True)
    parser.add_argument("--out-dir", default="portal/data/weekly")
    parser.add_argument("--generated-at", default=date.today().isoformat())
    args = parser.parse_args()

    workbook = Path(args.workbook)
    wb = load_workbook(workbook, read_only=True, data_only=True)
    # 2026-W31起正式sheet名统一为“集团-客户动态”；兼容读取历史底表旧名。
    if "集团-客户动态" in wb.sheetnames:
        customer_sheet = "集团-客户动态"
    elif "客户动态" in wb.sheetnames:
        customer_sheet = "客户动态"
    else:
        customer_sheet = "客户候选名单"
    customers = read_sheet(wb, customer_sheet, CUSTOMER_MAP)
    sources = read_sheet(wb, "信源明细", SOURCE_MAP)
    event_sheet = "集团-展会线索" if "集团-展会线索" in wb.sheetnames else "展会线索"
    events = read_sheet(wb, event_sheet, EVENT_MAP)
    tenders = read_sheet(wb, "招投标线索", TENDER_MAP)
    customers = [normalize_industry(row) for row in customers if row.get("week") == args.week]
    customer_ids_by_owner = defaultdict(list)
    for row in customers:
        owner = str(row.get("owner_company_brand") or "").strip()
        if owner:
            customer_ids_by_owner[owner].append(str(row.get("candidate_id") or ""))
    duplicate_owners = {
        owner: candidate_ids
        for owner, candidate_ids in customer_ids_by_owner.items()
        if len(candidate_ids) > 1
    }
    if duplicate_owners:
        duplicate_text = "; ".join(
            f"{owner}({', '.join(candidate_ids)})"
            for owner, candidate_ids in sorted(duplicate_owners.items())
        )
        raise ValueError(
            f"{args.week}客户动态必须按客户主体唯一；请先聚合同周重复企业：{duplicate_text}"
        )
    sources = [row for row in sources if row.get("week") == args.week]
    generated_day = datetime.strptime(args.generated_at, "%Y-%m-%d").date()
    # Excel keeps a 30-day collection pool; Web/weekly only shows the next 14 days.
    future_event_limit = generated_day + timedelta(days=14)
    events = [
        row for row in events
        if parse_sheet_date(row.get("date"))
        and generated_day <= parse_sheet_date(row.get("date")) <= future_event_limit
    ]
    events = sorted(events, key=lambda row: (row.get("date", ""), row.get("event_name", "")))
    tenders = [row for row in tenders if row.get("week") == args.week]
    tenders = filter_active_tenders(tenders, args.generated_at)

    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    weekly = build_weekly(customers, sources, events, tenders, args.week, args.generated_at, workbook)
    industry = build_industry_supply(weekly["leads_module_content"]["records"], args.week, args.generated_at, workbook)

    weekly_path = out_dir / f"weekly_leads_content_2026_{args.week}.json"
    industry_path = out_dir / f"industry_brief_supply_2026_{args.week}.json"
    weekly_path.write_text(json.dumps(weekly, ensure_ascii=False, indent=2), encoding="utf-8")
    industry_path.write_text(json.dumps(industry, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"wrote {weekly_path}")
    print(f"wrote {industry_path}")


if __name__ == "__main__":
    main()
